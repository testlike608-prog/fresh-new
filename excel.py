from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd
import os
import io
import shutil
from datetime import datetime
from PIL import Image as PILImage

import serial


def _base_dir():
    """المجلد المرجعي اللي بنحسب منه المسارات النسبية."""
    return Path(__file__).parent


def _resolve_folder(folder):
    """
    تحوّل اسم/مسار فولدر لـ Path كامل.
    - لو المسار مطلق (absolute) بيترجع زي ما هو.
    - لو نسبي بيتحسب جنب ملف الكود.
    """
    p = Path(str(folder).strip())
    if not p.is_absolute():
        p = _base_dir() / p
    return p


def _get_images_folder():
    """ترجع فولدر صور النتيجة من config (قابل للتعديل من الـ GUI)."""
    folder = "results"
    try:
        from config import config as _cfg
        folder = _cfg.get("result_images_folder", "results") or "results"
    except Exception:
        pass
    return _resolve_folder(folder)


def get_image_path(image_name, folder_name=None):
    """
    فانكشن بتاخد اسم الصورة وترجع الباث الكامل.
    لو folder_name = None بتاخد الفولدر من config (قابل للتعديل من الـ GUI).
    """
    # تحديد مسار فولدر الصور
    if folder_name is None:
        base = _get_images_folder()
    else:
        base = _resolve_folder(folder_name)

    # دمج المسار مع اسم الصورة
    image_path = base / image_name

    # التأكد إذا كانت الصورة موجودة فعلاً (اختياري لكن مفيد)
    if image_path.exists():
        return str(image_path.resolve())
    else:
        return f"خطأ: الصورة '{image_name}' غير موجودة في فولدر '{base}'"


def _copy_result_images_to_backups(image_names):
    """
    ينسخ صور النتيجة من الفولدر الأساسي لكل فولدرات الـ backup
    المحددة في config (result_images_backup_folders).
    """
    try:
        from config import config as _cfg
        backups = _cfg.get("result_images_backup_folders", []) or []
    except Exception:
        return

    if not backups:
        return

    src_folder = _get_images_folder()
    for folder in backups:
        folder = str(folder).strip()
        if not folder:
            continue
        dest = _resolve_folder(folder)
        try:
            dest.mkdir(parents=True, exist_ok=True)
            for name in image_names:
                src = src_folder / name
                if src.exists():
                    shutil.copy2(src, dest / name)
        except Exception as e:
            print(f"تعذّر نسخ صور النتيجة إلى '{dest}': {e}")
""""
# --- أمثلة للاستخدام ---

# 1. لو الصورة اسمها pic1.png وفي فولدر اسمه images
print(get_image_path("pic1.png"))

print(get_image_path("logo.jpg", folder_name="assets"))

# 3. استخدام متغير لاسم الصورة
my_photo = "profile.jpeg"
full_link = get_image_path(my_photo)
print(full_link)
"""
def result_reporting(ID, result, file_path=None):
    """
    يكتب نتيجة الفحص في Excel ويضم صور النتيجة مباشرة جوّا الخلايا.

    التنسيق:
        A: ID  |  B: Serial  |  C: Result  |  D: Timestamp  |  E,F,G...: صور

    عدد الصور = vision_test_count من config (الرقم اللي بتحدده في الـ GUI).
    الصور بتتضام كصور حقيقية في الخلايا مش كـ path نصي.
    """
    # ─── مسار الملف ───
    if file_path is None:
        try:
            from config import config as _cfg
            file_path = _cfg.get("results_report_file", "results_report.xlsx")
        except Exception:
            file_path = "results_report.xlsx"

    # ─── عدد الصور من config (نفس رقم vision_test_count في الـ GUI) ───
    try:
        from config import config as _cfg
        img_count = int(_cfg.get("vision_test_count", 4))
    except Exception:
        img_count = 4
    img_count = max(1, min(30, img_count))

    # ─── مقاسات الصورة داخل Excel ───
    IMG_W   = 180    # عرض الصورة (pixel)
    IMG_H   = 135    # ارتفاع الصورة (pixel)
    COL_W   = 27     # عرض عمود الصورة (وحدة Excel characters ≈ 7px لكل char)
    ROW_H   = 105    # ارتفاع صف الصورة (points — 1pt ≈ 1.33px)

    # ─── أعمدة الهيدر ───
    FIXED_HEADERS = ["ID",  "Result", "Timestamp"]
    IMG_HEADERS   = [f"Image {i + 1}" for i in range(img_count)]
    ALL_HEADERS   = FIXED_HEADERS + IMG_HEADERS

    # ─── تجميع أسماء ومسارات الصور ───
    images_folder = _get_images_folder()
    image_names   = [f"{ID}_{i}.jpg" for i in range(img_count)]
    image_paths   = [images_folder / name for name in image_names]
    _copy_result_images_to_backups(image_names)

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ─── فتح أو إنشاء الـ workbook ───
    is_new_file = not os.path.exists(file_path)

    if is_new_file:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        _write_header(ws, ALL_HEADERS, img_count, COL_W)
        next_row = 2
    else:
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            print(f"[Excel] تعذّر فتح {file_path}: {e} — هيتعمل ملف جديد")
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            _write_header(ws, ALL_HEADERS, img_count, COL_W)

        # لو الملف القديم فيه columns مختلفة (مسارات بدل صور) نعمل migrate
        if _needs_migration(ws, img_count):
            backup = _backup_old_file(file_path)
            print(f"[Excel] تنسيق قديم — تم نسخ الملف القديم إلى: {backup}")
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            _write_header(ws, ALL_HEADERS, img_count, COL_W)

        next_row = _find_next_row(ws)

    # ─── كتابة البيانات النصية ───
    ws.cell(row=next_row, column=1, value=str(ID))
    result_cell = ws.cell(row=next_row, column=3, value=str(result))
    # لون النتيجة: أخضر للـ PASS، أحمر للـ FAIL
    _color_result_cell(result_cell, str(result))
    ws.cell(row=next_row, column=4, value=current_time)

    # ─── تضمين الصور في الخلايا ───
    ws.row_dimensions[next_row].height = ROW_H

    imgs_embedded = 0
    # ── نحتفظ بالـ BytesIO objects هنا عشان openpyxl بيقرأهم وقت wb.save()
    # لو اتحذفوا من الميموري قبل كده بيظهر أول صورة بس ──────────────────
    _alive_bufs = []

    for i, img_path in enumerate(image_paths):
        col_idx    = 5 + i
        col_letter = get_column_letter(col_idx)

        if not img_path.exists():
            print(f"[Excel] صورة {i+1} غير موجودة: {img_path}")
            ws.cell(row=next_row, column=col_idx, value="—")
            continue

        try:
            # نعمل PIL resize هنا مباشرة ونحتفظ بالـ buf في القائمة
            pil = PILImage.open(str(img_path)).convert("RGB")
            pil.thumbnail((IMG_W, IMG_H), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format="JPEG")
            buf.seek(0)
            _alive_bufs.append(buf)          # منع garbage collection

            xl_img = XLImage(buf)
            xl_img.width  = IMG_W
            xl_img.height = IMG_H
            xl_img.anchor = f"{col_letter}{next_row}"
            ws.add_image(xl_img)
            imgs_embedded += 1
            print(f"[Excel] صورة {i+1} جاهزة: {img_path.name}")
        except Exception as e:
            ws.cell(row=next_row, column=col_idx, value=f"خطأ في الصورة: {e}")
            print(f"[Excel] خطأ في تضمين {img_path.name}: {e}")

    # ─── حفظ الملف (الـ buffers لازم تكون موجودة لحد هنا) ───
    try:
        wb.save(file_path)
        print(f"[Excel] ✅ تمت إضافة البيانات — صف {next_row} | صور: {imgs_embedded}/{img_count}")
    except Exception as e:
        print(f"[Excel] ❌ خطأ في حفظ {file_path}: {e}")
    finally:
        _alive_bufs.clear()   # الآن آمن نحذفهم


# ════════════════════════════════════════════════════════════════════
#                     Helper functions (Excel)
# ════════════════════════════════════════════════════════════════════

def _write_header(ws, headers, img_count, col_w):
    """يكتب صف الهيدر ويضبط العروض."""
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # عروض الأعمدة الثابتة
    ws.column_dimensions["A"].width = 16   # ID
    ws.column_dimensions["B"].width = 22   # Serial
    ws.column_dimensions["C"].width = 12   # Result
    ws.column_dimensions["D"].width = 20   # Timestamp

    # عروض أعمدة الصور
    for i in range(img_count):
        ws.column_dimensions[get_column_letter(5 + i)].width = col_w

    ws.row_dimensions[1].height = 22


def _color_result_cell(cell, result: str):
    """يلوّن خلية النتيجة: أخضر للـ PASS، أحمر للـ FAIL."""
    r = result.upper()
    if "PASS" in r or "نجح" in r:
        cell.fill = PatternFill("solid", fgColor="D4EDDA")
        cell.font = Font(color="155724", bold=True)
    elif "FAIL" in r or "فشل" in r:
        cell.fill = PatternFill("solid", fgColor="F8D7DA")
        cell.font = Font(color="721C24", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _prepare_xl_image(img_path: Path, w: int, h: int) -> XLImage:
    """يفتح الصورة، يصغّرها للمقاس المطلوب، ويرجع XLImage جاهز للإضافة."""
    pil = PILImage.open(str(img_path)).convert("RGB")
    pil.thumbnail((w, h), PILImage.LANCZOS)

    buf = io.BytesIO()
    pil.save(buf, format="JPEG")
    buf.seek(0)

    xl = XLImage(buf)
    xl.width  = w
    xl.height = h
    return xl


def _needs_migration(ws, img_count: int) -> bool:
    """
    بيرجع True لو الملف القديم فيه columns بـ 'image path' بدل صور حقيقية.
    أو لو عدد أعمدة الصور مش مطابق للـ img_count الحالي.
    """
    try:
        # نقرأ الهيدر من الصف الأول
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [str(h).lower() if h else "" for h in headers]
        # لو فيه "image path" في أي عمود → format قديم
        if any("image path" in h for h in headers):
            return True
        # لو عدد أعمدة الصور مختلف
        img_cols = [h for h in headers if h.startswith("image ")]
        if img_cols and len(img_cols) != img_count:
            return True
    except Exception:
        pass
    return False


def _find_next_row(ws) -> int:
    """يرجع رقم أول صف فاضي بعد البيانات الحالية."""
    return ws.max_row + 1


def _backup_old_file(file_path: str) -> str:
    """ينسخ الملف القديم بـ timestamp في اسمه قبل الـ migration."""
    p    = Path(file_path)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = p.parent / f"{p.stem}_backup_{ts}{p.suffix}"
    shutil.copy2(file_path, dest)
    return str(dest)

# --- مثال على الاستخدام في مشروعك ---
# يمكنك استدعاء هذه الفانكشن في كل مرة ينتهي فيها النظام من فحص قطعة معينة
                        #ID 
def get_model_value(input_string , file_path, search_column):
    try:
        # 1. استخراج آخر 3 حروف من النص المدخل
        # هذه الخطوة مفيدة جداً في مشاريعك البرمجية للتعامل مع الأكواد المختصرة
        suffix = input_string[-3:]
        
        # 2. قراءة ملف الإكسيل
        df = pd.read_excel(file_path)
        
        # 3. البحث عن الصف الذي يحتوي على الـ suffix في العمود المحدد
        # نستخدم .astype(str) لضمان مطابقة النصوص حتى لو كانت البيانات في الإكسيل أرقاماً
        match = df[df[search_column].astype(str) == suffix]
        
        # 4. التأكد من وجود نتائج واستخراج القيمة من عمود 'model'
        if not match.empty:
            # الوصول لعمود 'model' في أول صف مطابق نتيجه البحث
            result_value = match.iloc[0]['model']
            return result_value
        else:
            return "لم يتم العثور على القيمة المطلوبة"
            
    except KeyError:
        return "خطأ: تأكد من صحة أسماء الأعمدة (العمود المراد البحث فيه أو عمود model)"
    except Exception as e:
        return f"حدث خطأ: {e}"

# --- تجربة الكود ---
# لنفترض أن النص المدخل ينتهي بـ "B01" وتريد البحث عنه في عمود اسمه "Code"
# ليجلب لك القيمة المقابلة في عمود "model"
if __name__ == "__main__":
    # كود الاختبار اللي كان بيتنفذ على الـ import — الآن جواه __main__ guard
    try:
        wb = load_workbook('test.xlsx')
        sheet = wb.active
        sheet['A1'] = "Result"
        sheet['A1'].font = Font(bold=True, color="FF0000")
        wb.save('test_styled.xlsx')
    except Exception as e:
        print(f"Test workbook step skipped: {e}")

    final_value = get_model_value("Project_B01", "production_data.xlsx", "Code")
    print(f"Model: {final_value}")
